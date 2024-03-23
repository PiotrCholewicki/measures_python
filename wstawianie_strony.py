import openpyxl as xl
import random
from copy import copy
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Font
path1 = 'C:\\Users\\Piotr\\Desktop\\mieszkania\\szablon.xlsx' #ZMIEN LOKALIZACJE
path2 = 'C:\\Users\\Piotr\\Desktop\\mieszkania\\test.xlsx' #ZMIEN LOKALIZACJE

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)

ws2.column_dimensions['B'].width = 41


class Strona:
    def __init__(self, gk, gl, ob_1, ob_3, bkz, bkk, blz, blk, roz, wts): #ob - obwody, gk - gniazdka w kuchni, gl - gniazdka w lazience,
        self.gk = gk
        self.gl = gl
        self.ob_1 = ob_1
        self.ob_3 = ob_3
        self.bkz = bkz
        self.bkk = bkk
        self.blz = blz
        self.blk = blk
        self.roz = roz
        self.wts = wts

# Funkcja do kopiowania wierszy wraz z formatowaniem i dodawaniem podziału strony
def copy_rows(source_sheet, target_sheet, start_row, end_row, offset, nr_klatki, nr_mieszkania):

    for row in source_sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            target_cell = target_sheet.cell(row=cell.row + offset, column=cell.column, value=cell.value)
            target_cell.font = copy(cell.font)
            target_cell.border = copy(cell.border)
            target_cell.fill = copy(cell.fill)
            target_cell.number_format = copy(cell.number_format)
            target_cell.protection = copy(cell.protection)
            target_cell.alignment = copy(cell.alignment)
            
    target_sheet.cell(row=2 + offset, column=7, value=nr_klatki)
    target_sheet.cell(row=7 + offset, column=9, value=nr_klatki)
    target_sheet.cell(row=2 + offset, column=8, value=nr_mieszkania)
    target_sheet.cell(row=7 + offset, column=10, value=nr_mieszkania)

    ws = target_sheet
    row_number = end_row+offset  # the row that you want to insert page break
    page_break = Break(id=row_number)  # create Break obj
    ws.row_breaks.append(page_break)  # insert page break
    ws.col_breaks.append
#funkcja uzupelniajaca strony na podstawie danych wprowadzonych przez użytkownika
def uzupelnij_strony(strona, offset):
    target_sheet = ws2
    index = 0
    index_ob = 0

    for i in range(strona.gk):
        rand = random.random()*0.35+0.9
        target_sheet.cell(row=13 + offset + index, column=1, value=index+1)
        target_sheet.cell(row=13 + offset + index, column=2, value='Gn pt 2P+Z,16A,250V kuchnia ')
        if not(strona.wts):
            target_sheet.cell(row=13 + offset + index, column=3, value='B  16A  S   ')
            target_sheet.cell(row=13 + offset + index, column=5, value=80)
            target_sheet.cell(row=13 + offset + index, column=9, value=rand*80)
        else:
            target_sheet.cell(row=13 + offset + index, column=3, value='WTs  16A  ')
            target_sheet.cell(row=13 + offset + index, column=5, value=40)
            target_sheet.cell(row=13 + offset + index, column=9, value=rand*40)
        target_sheet.cell(row=13 + offset + index, column=7, value=rand) 
        target_sheet.cell(row=13 + offset + index, column=11, value='pozytywny') 
        index+=1
    for i in range(strona.gl):
        rand = random.random()*0.35+0.9
        target_sheet.cell(row=13 + offset + index, column=1, value=index+1)
        target_sheet.cell(row=13 + offset + index, column=2, value='Gn nt hermet 2P+Z,16A,250V łazienka ')
        if not(strona.wts):
            target_sheet.cell(row=13 + offset + index, column=3, value='B  16A  S   ')
            target_sheet.cell(row=13 + offset + index, column=5, value=80)
            target_sheet.cell(row=13 + offset + index, column=9, value=rand*80)
        else:
            target_sheet.cell(row=13 + offset + index, column=3, value='WTs  16A  ')
            target_sheet.cell(row=13 + offset + index, column=5, value=40)
            target_sheet.cell(row=13 + offset + index, column=9, value=rand*40)
        target_sheet.cell(row=13 + offset + index, column=7, value=rand) 
        target_sheet.cell(row=13 + offset + index, column=11, value='pozytywny') 
        index+=1

    for i in range(ob_1): #1-fazowe
        #w 15% przypadkow uzupelnij pola obwodow o wartosci automatycznie
        rand_sel = random.random()
        target_sheet.cell(row=49 + offset + index_ob, column=1, value=index_ob+1)
        target_sheet.cell(row=49 + offset + index_ob, column=2, value='Obwod 1-fazowy ')
        if(rand_sel < 0.85):
            target_sheet.cell(row=49 + offset + index_ob, column=6, value='> 1MΩ ')
        else:
            rand = random.randrange(850,1450)
            target_sheet.cell(row=49 + offset + index_ob, column=6, value=rand)
        index_ob += 1
    for i in range(ob_3): #3-fazowe
        rand_sel = random.random()
        target_sheet.cell(row=49 + offset + index_ob, column=1, value=index_ob+1)
        target_sheet.cell(row=49 + offset + index_ob, column=2, value='Obwod 3-fazowy ')
        if(rand_sel < 0.85):
            for i in range(3, 13):
                target_sheet.cell(row=49 + offset + index_ob, column=i, value='> 1MΩ ')
        else:
            for i in range(3, 13):
                rand = random.randrange(850,1450)
                target_sheet.cell(row=49 + offset + index_ob, column=i, value=rand)
        index_ob += 1


    #dodawanie roznicowki 
    if(strona.roz):
        target_sheet.cell(row=38 + offset, column=2, value='Wyłącznik  różnicowo-prądowy :  pomiar  prądu  wyłączenia :  wynik  prawidłowy ,  wartość  prądu  zawarta  między  15  m A     oraz   30  m A.  Działanie  na  przycisk  TEST-reakcja  prawidłowa-nastąpiło  wyłączenie.   Wyłacznik  nadaje  się  do  eksploatacji.')

    #błędy wprowadzone przez uzytkownika:
    index_bk = 0
    index_bl = gk
    bold_font = Font(name = "Arial", bold=True)
    tab_bz = []
    tab_bk = []
    #bledy zerowania w kuchni
    for i in range(strona.bkz):
        target_sheet.cell(row=13 + offset + index_bk, column=7, value='NIE').font = bold_font
        target_sheet.cell(row=13 + offset + index_bk, column=9, value='NIE').font = bold_font
        target_sheet.cell(row=13 + offset + index_bk, column=11, value='NEGATYWNY').font = bold_font
        index_bk += 1
        tab_bz.append(index_bk)
    #bledy kolkow w kuchni
    for i in range(strona.bkk):
        target_sheet.cell(row=13 + offset + index_bk, column=7, value='NIE').font = bold_font
        target_sheet.cell(row=13 + offset + index_bk, column=9, value='NIE').font = bold_font
        target_sheet.cell(row=13 + offset + index_bk, column=11, value='NEGATYWNY').font = bold_font
        index_bk += 1
        tab_bk.append(index_bk)
    #bledy zerowania w lazience
    for i in range(strona.blz):
        target_sheet.cell(row=13 + offset + index_bl, column=7, value='NIE').font = bold_font
        target_sheet.cell(row=13 + offset + index_bl, column=9, value='NIE').font = bold_font
        target_sheet.cell(row=13 + offset + index_bl, column=11, value='NEGATYWNY').font = bold_font
        index_bl += 1
        tab_bz.append(index_bl)
    #bledy kolkow w lazience
    for i in range(strona.blk):
        target_sheet.cell(row=13 + offset + index_bl, column=7, value='NIE').font = bold_font
        target_sheet.cell(row=13 + offset + index_bl, column=9, value='NIE').font = bold_font
        target_sheet.cell(row=13 + offset + index_bl, column=11, value='NEGATYWNY').font = bold_font
        index_bl += 1
        tab_bk.append(index_bl)
    #wypisywanie jako zachowana, Z WYJ. PKT.
    bold_font = Font(name = "Arial", bold=True, italic=True)
    tab_combined = tab_bk + tab_bz
    tab_combined.sort()
    combined_string = ','.join([str(value) for value in tab_combined])
    string_bz = ""
    string_bk = ""
    if(strona.bkz + strona.blz > 0):
        string_bz = "BRAK OCHRONY PORAZENIOWEJ DLA PKT. "+','.join([str(value) for value in tab_bz])
    if(strona.bkk + strona.blk > 0):
        string_bk = " BRAK KOŁKÓW OCHRONNYCH DLA PKT. "+','.join([str(value) for value in tab_bk])
    napis_dol = string_bz+string_bk 
    if(strona.bkz + strona.bkk + strona.blz + strona.blk > 0): #suma wieksza od zera, blad wystapil
        target_sheet.cell(row=43 + offset, column=8, value='zachowana Z WYJ. PKT. '+combined_string).font = bold_font
        target_sheet.cell(row=71 + offset, column=2, value=napis_dol).font = bold_font

    
#funkcja która przyjmuje input typu "3+2", zamienia to na 3 obwody 1 fazowe, i 2 obwody 3-fazowe, 
#zwraca tablice w ktorej tab[0] to 1-fazowe, tab[1] to 3 fazowe

def podaj_obwody(wyrazenie):
    tab = [0,0]
    i = 0
    for _ in wyrazenie:
        if wyrazenie[i] == '+':
            break
        else:
            i+=1
    tab[0] = int(wyrazenie[0:i])
    tab[1] = wyrazenie[i+1:]
    if(tab[1] ==''):
        tab[1] = 0
    tab[1] = int(tab[1])
    return tab

i = 0
nr_klatki = 0
nr_mieszkania = 0
while True:
    if(nr_klatki == 0 or nr_mieszkania == 0):
        print("Podaj wartości dla pierwszej strony: ")
    else:
        print("\nOstatnio utworzono: "+nr_klatki+"/"+nr_mieszkania)
    nr_klatki = input("Podaj nr klatki lub napisz stop ")
    if(nr_klatki == "STOP" or nr_klatki == "stop" or nr_klatki == "nie" or nr_klatki == "NIE" or nr_klatki == "n" or nr_klatki == "N"):
        break
    nr_mieszkania = input("Podaj nr mieszkania ")
    copy_rows(ws1, ws2, 1, ws1.max_row, ws1.max_row * i, nr_klatki, nr_mieszkania)
    ob = input("Podaj ilosc obwodow: ")
    #korzystajac z funkcji podaj_obwody przepisywane są wartości obwodow 1 i 3 fazowych do dwoch osobnych zmiennych
    tab = podaj_obwody(ob)
    ob_1 = tab[0] 
    ob_3 = tab[1]
    gk = int(input("Podaj ilosc gniazdek w kuchni "))
    gl = int(input("Podaj ilosc gniazdek w łazience "))
    check = input("Błędy/Różnicówka/WTS? ")
    check = check.upper()
    if(check == "T" or check == "TAK"):
        b_k_z = int(input("Kuchnia zerowanie: "))
        b_k_k = int(input("Kuchnia kołki: "))
        b_l_z = int(input("Łazienka zerowanie: "))
        b_l_k = int(input("Łazienka kołki: "))
        roznicowka = input("Różnicówka?(t/n): ")
        roznicowka = roznicowka.upper()
        if(roznicowka == "T" or roznicowka == "TAK"):
            roznicowka = True
        else:
            roznicowka = False
        wts = input("WTS?(t/n): ")
        wts = wts.upper()
        if(wts == "T" or wts == "TAK"):
            wts = True
        else:
            wts = False
        strona = Strona(gk, gl, ob_1, ob_3, b_k_z, b_k_k, b_l_z, b_l_k, roznicowka, wts)
    elif(check == "NIE" or check == "N" or check == ""):
        strona = Strona(gk, gl, ob_1, ob_3, 0, 0, 0, 0, False, False)
    else: #dla napisow check = "brw" albo "br", gdzie r to roznicowka, w to wts. Jesli 'w' wystepuje w lancuchu to znaczy ze dodac wts, jesli 'r' to roznicowke 
        roznicowka = False
        wts = False
        b_k_z = 0
        b_k_k = 0
        b_l_z = 0
        b_l_k = 0
        for i2 in check:
            if(i2 == "B"):
                b_k_z = int(input("Kuchnia zerowanie: "))
                b_k_k = int(input("Kuchnia kołki: "))
                b_l_z = int(input("Łazienka zerowanie: "))
                b_l_k = int(input("Łazienka kołki: "))
            elif(i2 == "R"):
                roznicowka = True
            elif(i2 == "W"):
                wts = True
        strona = Strona(gk, gl, ob_1, ob_3, b_k_z, b_k_k, b_l_z, b_l_k, roznicowka, wts)

    uzupelnij_strony(strona, ws1.max_row*i)
    # Kopiowanie scalonych komórek - kolejne strony
    for merged_cell_range in ws1.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_cell_range.min_row, merged_cell_range.min_col, merged_cell_range.max_row, merged_cell_range.max_col
        ws2.merge_cells(start_row=min_row + ws1.max_row * i, start_column=min_col, end_row=max_row + ws1.max_row * i, end_column=max_col)
    i += 1

    wb2.save(path2)
